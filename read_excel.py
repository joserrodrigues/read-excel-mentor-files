# -*- coding: utf-8 -*-
# ************************************************************
# Autor.......: Vladmir Cruz
# Data........: 06 de Junho de 2022
# Arquivo.....:
# Descricao...: Programa que le e escreve em arquivos Excel
# ************************************************************
# pip3 install requests openpyxl python-dotenv

# Imports
from fileinput import filename
import openpyxl # biblioteca que faz a leitura e gravação em arquivos Excel
from openpyxl.worksheet.table import Table, TableStyleInfo
import os       # biblioteca que acessa funções do sistema operacional
import sys # biblioteca que acessa funções do sistema base, nesse caso nos retorna o sistema operacional
import shutil # biblioteca para manipular arquivos
import logging
import json
import requests
import os
from dotenv import load_dotenv
import datetime
import warnings
import sys

def checkSendToBI():
    # total arguments
    n = len(sys.argv)

    if (len(sys.argv) == 1):
        return False

    for i in range(1, n):
        if(sys.argv[i] == '-s'): 
            return True


def createAppFiles(log_file_name, isMac):
    if os.path.exists(log_file_name):
        os.remove(log_file_name)    

    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


    load_dotenv()
    logging.basicConfig(filename=log_file_name, level=logging.DEBUG)        

    readFolder = var_loc + '\\lidos\\'
    resultFolder = var_loc + '\\resultado\\'
    if(isMac):
        readFolder = var_loc + '/lidos/'
        resultFolder = var_loc + '/resultado/'

    if(os.path.isdir(readFolder) == False):
        os.mkdir(readFolder) # diretorio que recebera os arquivos que foram lidos
    if(os.path.isdir(resultFolder) == False):  # se não existe o diretório resultado
        os.mkdir(resultFolder) # cria diretorio que contem o arquivo resultante

    if os.path.exists(getResultFile(isMac)):
        os.remove(getResultFile(isMac))

def sendImportantMessage(message):
    print(message)
    logging.debug(message)

def getResultFile(isMac):    
    if(isMac):
        return var_loc + '/resultado/' + 'resultado.xlsx'        
    else:
        return var_loc + '\\resultado\\' + 'resultado.xlsx'


def createResult(var_loc, isMac):
    wb = openpyxl.Workbook()    
    mainWks = wb.active
    
    # Preenchendo o cabeçalho da planilha resultante
    mainWks.cell(row=1, column=1).value = 'Nome do Arquivo'
    mainWks.cell(row=1, column=2).value = 'Mes'
    mainWks.cell(row=1, column=3).value = 'Ano'
    mainWks.cell(row=1, column=4).value = 'Nome do Professor'
    mainWks.cell(row=1, column=5).value = 'Contrato do Professor'
    mainWks.cell(row=1, column=6).value = 'Curso'
    mainWks.cell(row=1, column=7).value = 'Turma'
    mainWks.cell(row=1, column=8).value = 'Fase'
    mainWks.cell(row=1, column=9).value = 'Atividade'
    mainWks.cell(row=1, column=10).value = 'Data'
    mainWks.cell(row=1, column=11).value = 'Quantidade de Horas'
    mainWks.cell(row=1, column=12).value = 'Observacao'

    for courses in array_courses:
        wks = wb.create_sheet(title=courses)
        wks.cell(row=1, column=1).value = 'EMPRESA'
        wks.cell(row=1, column=2).value = 'CAMPUS'
        wks.cell(row=1, column=3).value = 'TIPO'
        wks.cell(row=1, column=4).value = 'UNIDADE / DEPARTAMENTO'
        wks.cell(row=1, column=5).value = 'SEGMENTO'
        wks.cell(row=1, column=6).value = 'DETALHE (CURSO)'
        wks.cell(row=1, column=7).value = 'Tipo Contrato'
        wks.cell(row=1, column=8).value = 'Nº Matrícula (CLT) | Contato (PJ/RPA)'
        wks.cell(row=1, column=9).value = 'Responsável - Nome Completo'
        wks.cell(row=1, column=10).value = 'Modalidade'
        wks.cell(row=1, column=11).value = 'Turma'
        wks.cell(row=1, column=12).value = 'Disciplina / Fase'
        wks.cell(row=1, column=13).value = 'Num. Cap.'
        wks.cell(row=1, column=14).value = 'Capítulo'
        wks.cell(row=1, column=15).value = 'DATA (Entrega / Correção)'
        wks.cell(row=1, column=16).value = 'CONTROLE INTERNO - LAUDAS / HORAS (não usar)'
        wks.cell(row=1, column=17).value = 'Núm. págs / horas'
        wks.cell(row=1, column=18).value = 'Data de cadastro (interno)'        

    wb.save(getResultFile(isMac))  # salva o workbook

def getRightWorksheet(workbook):
    ind = 0
    for worksheet in workbook.worksheets:
        if ("Preencher" in str(worksheet)):
            return ind
        ind += 1
    return -1

def checkCellInfo(worksheet, row, column, info_waited):
    aux_info = worksheet.cell(row, column).value
    if(not isinstance(aux_info, str) or info_waited.lower() not in aux_info.lower()):
        sendImportantMessage("[ERROR] Planilha com formato inválido: " + info_waited +" != "+ str(aux_info) + " \n\n")
        return False
    return True


def getProfessorName(workbook):
    work_sheet_ind = getRightWorksheet(workbook)
    worksheet = workbook.worksheets[work_sheet_ind]
    var_prf = worksheet.cell(row=4, column=3).value # lendo o nome do professor
        
    if(str(var_prf) == "" or str(var_prf) == "None"):
        var_prf = worksheet.cell(row=4, column=1).value # lendo o nome do professor
        var_prf = var_prf.replace("Nome do professor: ", "")

    if(str(var_prf) == "" or str(var_prf) == "None"):
        var_prf = worksheet.cell(row=4, column=2).value # lendo o nome do professor
        var_prf = var_prf.replace("Nome do professor: ", "")        

    return var_prf


def checkWorkSheetPattern(workbook):

    work_sheet_ind = getRightWorksheet(workbook)
    if(work_sheet_ind == -1):
        sendImportantMessage("[ERROR] Planilha com formato inválido: Worksheet fora do padrao \n\n")
        return False    

    worksheet = workbook.worksheets[work_sheet_ind]

    if (not checkCellInfo(worksheet, 1, 1, "Apontamento de horas mensal")):
        return False
    if (not checkCellInfo(worksheet, 3, 1, "Mês")):
        return False
    if (not checkCellInfo(worksheet, 4, 1, "Nome do professor")):
        return False
    if (not checkCellInfo(worksheet, 5, 1, "Tipo de contrato")):
        return False
    if (not checkCellInfo(worksheet, 7, 1, "Curso")):
        return False
    if (not checkCellInfo(worksheet, 7, 2, "Turma")):
        return False
    if (not checkCellInfo(worksheet, 7, 3, "Fase")):
        return False
    if (not checkCellInfo(worksheet, 7, 4, "Tipo de atividade")):
        return False
    if (not checkCellInfo(worksheet, 7, 5, "Data")):
        return False
    if (not checkCellInfo(worksheet, 7, 6, "de horas")):
        return False
    if (not checkCellInfo(worksheet, 7, 7, "Observação")):
        return False
    
    var_prf = getProfessorName(workbook)

    if(str(var_prf) == "" or str(var_prf) == "None"):
        sendImportantMessage(
            "[ERROR] Planilha com formato inválido: Nome do Professor em branco \n\n")
        return False

    return True

def mountInfo(wb, array_send, line_controller, fileName, month, year, prof, contract, course, class_txt, phase, activity, 
                final_data, final_machine_data, hours, obs):
    line_send = {}
    line_send['nome_arquivo'] = str(fileName)
    line_send['Mes'] =str(month)
    line_send['Ano'] = str(year)
    line_send['Professor'] = str(prof)
    line_send['Contrato'] = str(contract)
    line_send['Curso'] =str(course)
    line_send['Turma'] = str(class_txt)
    line_send['Fase'] = str(phase)
    line_send['Atividade'] = str(activity)
    line_send['Data'] = str(final_machine_data)
    line_send['Quantidade'] = str(hours)
    line_send['Observacao'] = str(obs)

    array_send.append(line_send)

    wkLine = line_controller['all']

    work_sheet = wb['Sheet']

    work_sheet.cell(row=wkLine, column=1).value = "VSTP"
    work_sheet.cell(row=wkLine, column=2).value = "FIAP ON"
    work_sheet.cell(row=wkLine, column=3).value = "EDUCACIONAL / PROFESSORES"
    work_sheet.cell(row=wkLine, column=4).value = "MBA ON"
    work_sheet.cell(row=wkLine, column=5).value = "Especialização"
    work_sheet.cell(row=wkLine, column=6).value = str(course).upper()
    work_sheet.cell(row=wkLine, column=7).value = "**Copiar"
    work_sheet.cell(row=wkLine, column=8).value = "**Copiar"
    work_sheet.cell(row=wkLine, column=9).value = str(prof).upper()
    work_sheet.cell(row=wkLine, column=10).value = str(activity)
    work_sheet.cell(row=wkLine, column=11).value = str(class_txt).replace(" ", "")
    work_sheet.cell(row=wkLine, column=12).value = str(phase)
    work_sheet.cell(row=wkLine, column=13).value = ""
    work_sheet.cell(row=wkLine, column=14).value = ""
    work_sheet.cell(row=wkLine, column=15).value = str(final_data)
    work_sheet.cell(row=wkLine, column=16).value = "**Copiar"
    work_sheet.cell(row=wkLine, column=17).value = str(hours).replace(".", ",")
    work_sheet.cell(row=wkLine, column=18).value = ""
    
    wkLine = line_controller[course]
    work_sheet = wb[course]
    work_sheet.cell(row=wkLine, column=1).value = "VSTP"
    work_sheet.cell(row=wkLine, column=2).value = "FIAP ON"
    work_sheet.cell(row=wkLine, column=3).value = "EDUCACIONAL / PROFESSORES"
    work_sheet.cell(row=wkLine, column=4).value = "MBA"
    work_sheet.cell(row=wkLine, column=5).value = "Especialização"
    work_sheet.cell(row=wkLine, column=6).value = str(course).upper()
    work_sheet.cell(row=wkLine, column=7).value = "**Copiar"
    work_sheet.cell(row=wkLine, column=8).value = "**Copiar"
    work_sheet.cell(row=wkLine, column=9).value = str(prof).upper()
    work_sheet.cell(row=wkLine, column=10).value = str(activity)
    work_sheet.cell(row=wkLine, column=11).value = str(class_txt).replace(" ", "")
    work_sheet.cell(row=wkLine, column=12).value = str(phase)
    work_sheet.cell(row=wkLine, column=13).value = ""
    work_sheet.cell(row=wkLine, column=14).value = ""
    work_sheet.cell(row=wkLine, column=15).value = str(final_data)
    work_sheet.cell(row=wkLine, column=16).value = "**Copiar"
    work_sheet.cell(row=wkLine, column=17).value = str(hours).replace(".", ",")
    work_sheet.cell(row=wkLine, column=18).value = ""

    line_controller['all'] += 1
    line_controller[course] += 1

def sendPowerBI(array_send,wb):
    logging.debug("Array to Send")
    logging.debug(array_send)
    res = json.dumps(array_send)
    r = requests.post(os.environ['URL']+ os.environ['KEY'], data=res)

    logging.debug("Sending to Power BI")
    if(r.status_code == 200):
        sendImportantMessage("Sending to Power BI - Success " + str(r.status_code))
    else:
        sendImportantMessage("Sending to Power BI - Error Code " + str(r.status_code) + " \n Status:" + str(r.text))
        return False
    return True


# Variáveis
var_loc = os.getcwd() # caminho do diretório pai do script
var_loc_read = os.getcwd()+"/original"  # caminho do diretório de leitura
var_loc_done = os.getcwd()+"/lidos/"  # caminho do diretório de leitura
var_dir = os.listdir(var_loc_read) # acessando o diretório que contêm os arquivos
var_os = sys.platform # sistema operacional que o programa está rodando
var_flt = [] # lista de arquivos que serão lidos para o conteúdo 
var_mon = '' # mês que está sendo reportado
var_yea = '' # ano que está sendo reportado
var_prf = '' # nome do professor que está reportando
var_tct = '' # tipo de contrato do professor
var_crs = '' # curso que o professor está reportando
var_trm = '' # turma que o professor está reportando
var_fas = '' # fase do curso que o professor está reportando
var_atv = '' # tipo de atividade que o professor está reportando
var_dat = '' # data que ocorreu o evento sendo reportado
var_qhr = '' # quantidade de horas que o professor está reportando
var_obs = '' # observação que o professor possa fazer
var_naq = '' # variável para armazenar o nome do arquivo atual
var_col = 0 # um contador genérico para contar as colunas
var_lin = 0 # um contador genérico para contar as linhas
var_ctr = 0 # um contador genérico, afinal, todo programa precisa de um
var_wkl = 0 # um contador de linhas para o Workbook
line_controller = {}
array_courses = ['GTIO', 'AOJO', 'ASOO', 'ABDO', 'DTSO', 'BDTO', 'DGO', 'NGO', 'BIO', 'SGO', 'SCJO', 'STO'] 

isSendToBI = checkSendToBI()

isMac = False
# verifica se o sistema operacional é Linux ou MacOS
if(var_os == 'linux' or var_os == 'linux2' or var_os == 'darwin'):
    isMac = True

#cria os arquivos e pastas utilizadas no app
createAppFiles('reading.log', isMac)
info_message = "\n ------------------------------------------------------------------"
info_message += "\n  Iniciando a importacao "
info_message += "\n ------------------------------------------------------------------"
sendImportantMessage(info_message)

# verifica se os diretorios resultantes existem, se não, os cria
if(os.path.isfile(getResultFile(isMac)) == False):
    #Cria a planilha de resultado    
    createResult(var_loc, isMac)
    
#abre a planilha de resultado
var_wkb = openpyxl.load_workbook(getResultFile(isMac))
var_wks = var_wkb.active  # criando a sheet que receberá os dados


for files in var_dir: # lendo todos os arquivos que estão armazenados no diretório
    if(files.endswith('.xls') or files.endswith('.xlsx')):
        if(var_os == 'linux' or var_os == 'linux2' or var_os == 'darwin'): # verifica se o sistema operacional é Linux ou MacOS
            var_flt.append(var_loc_read + '/' + files) # grava o nome do arquivo no array de nomes
        elif var_os == 'win32': # verifica se o sistema operacional é Windows
            var_flt.append(var_loc_read + '\\' + files) # grava o nome do arquivo no array de nomes

#mount lines Controller
line_controller['all'] = 1
for course in array_courses:
    line_controller[course] = 2

total_process_number = 0
success_process_number = 0
error_process_number = 0
success_send_bi_number = 0
error_send_bi_number = 0
logging.debug("ARQUIVOS Excel a ler")
logging.debug(var_flt)
# Iterando sobre a lista de nomes de arquivos e seu conteúdo
for wb in var_flt:

    total_process_number +=1
    array_send = []
    var_naq = var_flt[var_ctr]
    is_reading_ok = False
    sendImportantMessage("Reading: " + wb)

    try:
        workbook = openpyxl.load_workbook(wb) # chamada que abre o arquivo desejado
        is_reading_ok = True
    except:
        sendImportantMessage("[ERROR] Erro ao abrir a planilha")

    if(is_reading_ok and checkWorkSheetPattern(workbook)):
        # definindo qual o caderno do Excel que será utilizado
        work_sheet_ind = getRightWorksheet(workbook)
        worksheet = workbook.worksheets[work_sheet_ind]

        var_mon = worksheet.cell(row=3, column=2).value # lendo o mês do relatório
        var_yea = worksheet.cell(row=3, column=6).value # lendo o ano do relatório
        var_prf = getProfessorName(workbook) # lendo o nome do professor
        var_tct = worksheet.cell(row=5, column=3).value # lendo o tipo de contrato do professor
        
        if(str(var_prf) == "" or str(var_prf) == "None"):
            var_prf = worksheet.cell(row=4, column=1).value # lendo o nome do professor
            var_prf = var_prf.replace("Nome do professor: ", "")

        var_lin = 8 # contador começa em 8 por ser a primeira linha que contém informações sobre as atividades
        var_col = 1 # contador começa em 1 por ser a primeira coluna que contém informações sobre as atividades
        while var_lin < 270: # iterando sobre a tabela com os dados das atividades reportadas        
            var_check_end = worksheet.cell(row=var_lin, column=6).value # lendo o curso que o professor indicou        
            if (isinstance(var_check_end, str) and "SUM" in var_check_end): # Checa se o cursor chegou no final da planilha
                var_lin = 300
            else:
                var_crs = worksheet.cell(row=var_lin, column=var_col).value # lendo o curso que o professor indicou
                if(var_crs is not None): # verifica se há um curso sendo reportado, em caso de em branco, pula
                    var_trm = worksheet.cell(row=var_lin, column=var_col + 1).value # lendo a turma
                    var_fas = worksheet.cell(row=var_lin, column=var_col + 2).value # lendo a fase
                    var_atv = worksheet.cell(row=var_lin, column=var_col + 3).value # lendo a atividade
                    var_dat = worksheet.cell(row=var_lin, column=var_col + 4).value # lendo a data da ocorrência
                    var_qhr = worksheet.cell(row=var_lin, column=var_col + 5).value # lendo a quantidade de horas indicada
                    var_obs = str(worksheet.cell(row=var_lin, column=var_col + 6).value) # lendo a observação que foi apontada
                    var_col = 1 # resetando o contador de colunas
                    var_lin = var_lin + 1 # pulando para a próxima linha

                    if(str(var_obs) == "None"):
                        var_obs = ""

                    #Trabalhando com a data 
                    final_var_data = ""
                    final_machine_data = ""
                    if (isinstance(var_dat, datetime.datetime)):
                        final_var_data = var_dat.strftime("%d/%m/%Y")
                        final_machine_data = var_dat.strftime("%Y-%m-%d")
                    else:
                        newDate = datetime.date.today().replace(day=1)
                        final_var_data = newDate.strftime("%d/%m/%Y")
                        final_machine_data = newDate.strftime("%Y-%m-%d")
                        var_obs += " Data Corrigida - Dado Enviado = " + str(var_dat) 

                    mountInfo(var_wkb, array_send, line_controller, var_naq, var_mon, var_yea, var_prf, var_tct, var_crs,
                            var_trm, var_fas, var_atv, final_var_data, final_machine_data, var_qhr, var_obs)
                else:
                    var_lin = var_lin + 1 # pulando para a próxima linha

        # print(array_send)
        if isSendToBI:
            if  sendPowerBI(array_send, wb): #Envia para o PowerBI
                success_send_bi_number +=1
            else:
                error_send_bi_number += 1
        array_send.clear() # Limpa o array
        
        # Move para a pasta de lidos
        str_filename = wb.replace(var_loc_read, "")
        shutil.move(wb, var_loc_done+str_filename)
        
        var_ctr = var_ctr + 1 # pula para a próxima planiha quando termina a iteração da atual
        success_process_number +=1
    else:
        error_process_number += 1
    print ("\n ------------------------------------------------------------------\n\n")

var_wkb.save(var_loc + '/resultado/' + 'resultado.xlsx')
 
final_message = "\n ------------------------------------------------------------------"
final_message += "\n Fim de Processamento "
final_message += "\n Arquivo no processo: " + str(total_process_number)
final_message +="\n Processado com sucesso: " + str(success_process_number)
final_message +="\n Enviado ao BI com sucesso: " + str(success_send_bi_number)
final_message +="\n Enviado ao BI com erro: " + str(error_send_bi_number)
final_message +="\n Processado com erro: " + str(error_process_number)
final_message +="\n ------------------------------------------------------------------\n\n"

if(not isSendToBI):
    final_message += "\n ATENCAO: AS INFORMACOES NAO FORAM ENVIADAS PARA O BI " 
sendImportantMessage(final_message)

